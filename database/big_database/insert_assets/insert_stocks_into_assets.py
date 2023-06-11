import os
from openpyxl import load_workbook
import sqlite3


def _insert_stocks():

    # Verbindung zur Datenbank herstellen
    conn = sqlite3.connect('database/database_db_files/database.db')
    cursor = conn.cursor()

    # Funktion zum Einfügen eines Assets in die Datenbank
    def insert_asset(ticker, name, sector, asset_type):
        cursor.execute('INSERT OR IGNORE INTO Assets (ticker, name, sector, type) VALUES (?, ?, ?, ?)',
                    (ticker, name, sector, asset_type))

    # Funktion zum Verarbeiten einer Excel-Datei und Hinzufügen der Daten in die Datenbank
    def process_excel_file(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active

        # Überspringe den Header und iteriere über die Zeilen der Excel-Datei
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ticker, company, sector = row[0], row[1], row[2]
            asset_type = 'stock'

            # Füge das Asset in die Datenbank ein
            insert_asset(ticker, company, sector, asset_type)

    # Ordnerpfad für die Excel-Dateien
    folder_path = 'data/stocks/~index_ticker_list/'

    # Durchlaufe alle Dateien im Ordner und Unterordnern
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                process_excel_file(file_path)

    # Änderungen in der Datenbank speichern und Verbindung schließen
    conn.commit()
    conn.close()

    print('Import abgeschlossen.')

_insert_stocks()
